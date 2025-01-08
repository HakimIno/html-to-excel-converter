import sys
import logging
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("Error: beautifulsoup4 is not installed. Please run: pip install beautifulsoup4")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("Error: pandas is not installed. Please run: pip install pandas")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please run: pip install openpyxl")
    sys.exit(1)

try:
    import cssutils
except ImportError:
    print("Error: cssutils is not installed. Please run: pip install cssutils")
    sys.exit(1)

# ตั้งค่า logging level สำหรับ cssutils
cssutils.log.setLevel(logging.CRITICAL)

class HTMLToExcelConverter:
    def __init__(self):
        cssutils.log.setLevel(logging.CRITICAL)
        
        self.default_font = Font(name='TH Sarabun New', size=10)
        self.default_alignment = Alignment(horizontal='center', vertical='center')
        self.default_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_font = Font(name='TH Sarabun New', size=10, bold=True)
        self.current_row = 1

    def get_cell_dimensions(self, cell):
        """Get rowspan and colspan of a cell"""
        rowspan = int(cell.get('rowspan', 1))
        colspan = int(cell.get('colspan', 1))
        return rowspan, colspan
    
    def process_cell(self, ws, cell, row, col, merged_cells):
        """Process a single cell and handle merging"""
        rowspan, colspan = self.get_cell_dimensions(cell)
        
        # Get cell content
        value = cell.get_text(strip=True)
        excel_cell = ws.cell(row=row, column=col)
        excel_cell.value = value
        
        # Apply styles
        styles = self.parse_style(cell)
        
        # Set default style for header cells
        if cell.name == 'th':
            styles['font-weight'] = 'bold'
        
        self.apply_cell_style(excel_cell, styles)
        
        # Handle merged cells
        if rowspan > 1 or colspan > 1:
            merge_range = f"{get_column_letter(col)}{row}:{get_column_letter(col+colspan-1)}{row+rowspan-1}"
            try:
                ws.merge_cells(merge_range)
                
                # Track merged areas
                for r in range(row, row + rowspan):
                    for c in range(col, col + colspan):
                        merged_cells.add((r, c))
            except:
                print(f"Warning: Could not merge cells {merge_range}")
        
        return colspan
    
    def parse_style(self, element):
        """Parse CSS styles from element"""
        styles = {}
        
        # Get inline styles
        if element.get('style'):
            try:
                style = cssutils.parseStyle(element['style'])
                for prop in style:
                    styles[prop.name] = prop.value
            except:
                pass
                
        # Get background color from HTML attributes
        if element.get('bgcolor'):
            styles['background-color'] = element['bgcolor']
            
        return styles

    def css_to_rgb(self, color):
        """Convert CSS color to RGB tuple"""
        if not color:
            return None
            
        try:
            if color.startswith('#'):
                # Handle hex colors
                color = color.lstrip('#')
                if len(color) == 3:
                    color = ''.join(c + c for c in color)
                return tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
                
            elif color.startswith('rgb'):
                # Handle rgb/rgba colors
                return tuple(map(int, re.findall(r'\d+', color)[:3]))
        except:
            pass
            
        return None

    def apply_cell_style(self, cell, styles):
        """Apply styles to Excel cell"""
        try:
            # Background color
            if 'background-color' in styles:
                rgb = self.css_to_rgb(styles['background-color'])
                if rgb:
                    cell.fill = PatternFill(
                        start_color=f'{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}',
                        end_color=f'{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}',
                        fill_type='solid'
                    )

            # Font
            font_props = {'name': 'TH Sarabun New', 'size': 10}  # Default font properties
            if 'color' in styles:
                rgb = self.css_to_rgb(styles['color'])
                if rgb:
                    font_props['color'] = f'{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'
            if 'font-size' in styles:
                try:
                    size_match = re.search(r'(\d+)', styles['font-size'])
                    if size_match:
                        size = float(size_match.group(1))
                        font_props['size'] = size
                except:
                    pass
            if 'font-weight' in styles:
                font_props['bold'] = styles['font-weight'] in ('bold', '700')
                
            cell.font = Font(**font_props)

            # Alignment
            align_props = {'wrap_text': True}
            if 'text-align' in styles:
                if styles['text-align'] == 'left':
                    align_props['horizontal'] = 'left'
                elif styles['text-align'] == 'right':
                    align_props['horizontal'] = 'right'
                else:
                    align_props['horizontal'] = 'center'
            else:
                align_props['horizontal'] = 'center'
                
            align_props['vertical'] = 'center'  # Default vertical alignment
            cell.alignment = Alignment(**align_props)

            # Border
            cell.border = self.default_border
        except Exception as e:
            print(f"Error applying style: {str(e)}")

    def process_table(self, ws, table, start_row):
        """Process a single table and return the next row number"""
        current_row = start_row
        max_col = 0
        merged_cells = set()
        header_rows = []
        footer_rows = []
        body_rows = []

        # แยก header, body, footer rows
        rows = table.find_all('tr')
        for row in rows:
            if row.find('th'):
                header_rows.append(row)
            elif row.parent and row.parent.name == 'tfoot':
                footer_rows.append(row)
            else:
                body_rows.append(row)

        # Process header rows
        for row in header_rows:
            current_col = 1
            for cell in row.find_all(['th']):
                while (current_row, current_col) in merged_cells:
                    current_col += 1

                value = cell.get_text(strip=True)
                excel_cell = ws.cell(row=current_row, column=current_col)
                excel_cell.value = value

                rowspan = int(cell.get('rowspan', 1))
                colspan = int(cell.get('colspan', 1))

                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=current_col,
                        end_row=current_row + rowspan - 1,
                        end_column=current_col + colspan - 1
                    )
                    for r in range(current_row, current_row + rowspan):
                        for c in range(current_col, current_col + colspan):
                            merged_cells.add((r, c))

                current_col += colspan
            max_col = max(max_col, current_col - 1)
            current_row += 1

        # Process body rows
        for row in body_rows:
            current_col = 1
            for cell in row.find_all(['td']):
                while (current_row, current_col) in merged_cells:
                    current_col += 1

                value = cell.get_text(strip=True)
                excel_cell = ws.cell(row=current_row, column=current_col)
                excel_cell.value = value
                
                styles = self.parse_style(cell)
                self.apply_cell_style(excel_cell, styles)

                rowspan = int(cell.get('rowspan', 1))
                colspan = int(cell.get('colspan', 1))

                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=current_col,
                        end_row=current_row + rowspan - 1,
                        end_column=current_col + colspan - 1
                    )
                    for r in range(current_row, current_row + rowspan):
                        for c in range(current_col, current_col + colspan):
                            merged_cells.add((r, c))

                current_col += colspan
            max_col = max(max_col, current_col - 1)
            current_row += 1

        # Process footer rows
        for row in footer_rows:
            current_col = 1
            for cell in row.find_all(['td']):
                while (current_row, current_col) in merged_cells:
                    current_col += 1

                value = cell.get_text(strip=True)
                excel_cell = ws.cell(row=current_row, column=current_col)
                excel_cell.value = value
                
                styles = self.parse_style(cell)
                self.apply_cell_style(excel_cell, styles)
                excel_cell.font = Font(name='TH Sarabun New', size=10, bold=True)

                rowspan = int(cell.get('rowspan', 1))
                colspan = int(cell.get('colspan', 1))

                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=current_col,
                        end_row=current_row + rowspan - 1,
                        end_column=current_col + colspan - 1
                    )
                    for r in range(current_row, current_row + rowspan):
                        for c in range(current_col, current_col + colspan):
                            merged_cells.add((r, c))

                current_col += colspan
            max_col = max(max_col, current_col - 1)
            current_row += 1

        return current_row + 1, max_col

    def convert(self, html_content, output):
        """Convert HTML to Excel"""
        try:
            # Split content by DOCTYPE to handle multiple HTML documents
            html_documents = html_content.split('<!DOCTYPE html>')
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            
            current_row = 1
            max_col = 13  # Fixed number of columns based on the table structure

            # Process each HTML document
            for doc in html_documents:
                if not doc.strip():
                    continue
                    
                # Add DOCTYPE back for proper parsing
                doc = '<!DOCTYPE html>' + doc
                soup = BeautifulSoup(doc, 'html.parser')
                
                # Process the timestamp header
                timestamp_header = soup.find('th', string=lambda x: x and 'Printed' in x)
                if timestamp_header:
                    cell = ws.cell(row=current_row, column=max_col-2)
                    cell.value = timestamp_header.get_text(strip=True)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.font = Font(name='TH Sarabun New', size=10)
                    current_row += 1
                
                # Add blank row after timestamp
                current_row += 1
                
                # Process company header
                company_table = soup.find('table', attrs={'style': lambda x: x and 'margin-bottom: 5px' in x})
                if company_table:
                    # Company name
                    company_name = company_table.find('th', string=lambda x: x and 'บริษัท' in x)
                    if company_name:
                        cell = ws.cell(row=current_row, column=1)
                        cell.value = company_name.get_text(strip=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='TH Sarabun New', size=10, bold=True)
                        ws.merge_cells(f'A{current_row}:M{current_row}')
                        current_row += 1
                    
                    # Date range
                    date_range = company_table.find('th', string=lambda x: x and 'ระหว่างวันที่' in x)
                    if date_range:
                        cell = ws.cell(row=current_row, column=1)
                        cell.value = date_range.get_text(strip=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='TH Sarabun New', size=10)
                        ws.merge_cells(f'A{current_row}:M{current_row}')
                        current_row += 1
                    
                    # Add blank row after company info
                    current_row += 1
                    
                    # Process employee info rows
                    for row in company_table.find_all('tr'):
                        cells = row.find_all(['th'])
                        if len(cells) < 2:  # Skip rows without proper structure
                            continue
                            
                        col = 1
                        for i, cell in enumerate(cells):
                            text = cell.get_text(strip=True)
                            if not text or text in ['บริษัท', 'ระหว่างวันที่']:  # Skip already processed headers
                                continue
                                
                            excel_cell = ws.cell(row=current_row, column=col)
                            excel_cell.value = text
                            excel_cell.font = Font(name='TH Sarabun New', size=10)
                            
                            # Set alignment
                            if ':' in text:  # Labels
                                excel_cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                excel_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                            
                            # Handle colspan
                            colspan = int(cell.get('colspan', 1))
                            if colspan > 1:
                                ws.merge_cells(
                                    start_row=current_row, 
                                    start_column=col,
                                    end_row=current_row,
                                    end_column=col + colspan - 1
                                )
                            col += colspan
                        
                        if col > 1:  # Only increment row if we processed any cells
                            current_row += 1
                
                # Add blank rows before main table
                current_row += 2
                
                # Process main table
                main_table = None
                for table in soup.find_all('table', recursive=True):
                    if table.find('tr', attrs={'style': lambda x: x and 'text-align: center' in x and 'font-size: 10px' in x}):
                        main_table = table
                        break
                
                if main_table:
                    # Process table headers first
                    header_row = main_table.find('tr', attrs={'style': lambda x: x and 'text-align: center' in x and 'font-size: 10px' in x})
                    if header_row:
                        col = 1
                        for th in header_row.find_all('th'):
                            cell = ws.cell(row=current_row, column=col)
                            cell.value = th.get_text(strip=True)
                            cell.font = Font(name='TH Sarabun New', size=10, bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
                            cell.border = self.default_border
                            
                            # Adjust row height for header
                            ws.row_dimensions[current_row].height = 30
                            
                            col += 1
                        current_row += 1
                    
                    # Process table body
                    for tr in main_table.find_all('tr', attrs={'style': lambda x: x and 'text-align: center' in x and 'font-size: 10px' in x}):
                        if tr == header_row:
                            continue
                            
                        col = 1
                        cells = tr.find_all(['td'])
                        if not cells:
                            continue
                            
                        # Check if row should have gray background
                        has_gray_bg = any('background-color: #f0f0f0' in cell.get('style', '') for cell in cells)
                        
                        for td in cells:
                            cell = ws.cell(row=current_row, column=col)
                            cell.value = td.get_text(strip=True)
                            cell.font = Font(name='TH Sarabun New', size=10)
                            
                            # Apply styles
                            styles = self.parse_style(td)
                            if 'text-align' in styles:
                                if styles['text-align'] == 'left':
                                    cell.alignment = Alignment(horizontal='left', vertical='center')
                                elif styles['text-align'] == 'right':
                                    cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Apply background color
                            if has_gray_bg:
                                cell.fill = PatternFill(start_color='f0f0f0', end_color='f0f0f0', fill_type='solid')
                            
                            cell.border = self.default_border
                            col += 1
                        current_row += 1
                
                # Process footer
                footer = soup.find('tfoot')
                if footer:
                    footer_row = footer.find('tr')
                    if footer_row:
                        cells = footer_row.find_all('td')
                        if len(cells) >= 3:
                            # First cell
                            cell = ws.cell(row=current_row, column=1)
                            cell.value = cells[0].get_text(strip=True)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.font = Font(name='TH Sarabun New', size=10)
                            ws.merge_cells(f'A{current_row}:D{current_row}')
                            
                            # Second cell
                            cell = ws.cell(row=current_row, column=5)
                            cell.value = cells[1].get_text(strip=True)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.font = Font(name='TH Sarabun New', size=10)
                            ws.merge_cells(f'E{current_row}:I{current_row}')
                            
                            # Third cell
                            cell = ws.cell(row=current_row, column=10)
                            cell.value = cells[2].get_text(strip=True)
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.font = Font(name='TH Sarabun New', size=10)
                            ws.merge_cells(f'J{current_row}:M{current_row}')
                
                # Add extra spacing between documents
                current_row += 3
            
            # Adjust column widths
            for col in range(1, max_col + 1):
                max_length = 0
                column = get_column_letter(col)
                
                for cell in ws[column]:
                    if cell.value:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            max_length = max(max_length, 30)
                            
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save workbook
            if isinstance(output, str):
                wb.save(output)
            else:
                wb.save(output)  # Save to buffer
            print(f"Successfully converted to Excel")
            
        except Exception as e:
            print(f"Error converting HTML to Excel: {str(e)}")
            raise

    @classmethod
    def convert_file(cls, input_path, output_path):
        """Convert HTML file to Excel file"""
        try:
            with open(input_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            converter = cls()
            converter.convert(html_content, output_path)
            
        except Exception as e:
            print(f"Error converting file: {str(e)}")
            raise

if __name__ == "__main__":
    # ตรวจสอบ command line arguments
    if len(sys.argv) < 2:
        print("กรุณาระบุชื่อไฟล์ HTML")
        print("วิธีใช้: python converter.py input.html [output.xlsx]")
        sys.exit(1)

    # รับชื่อไฟล์ input
    input_file = sys.argv[1]
    
    # กำหนดชื่อไฟล์ output (ถ้าไม่ระบุจะใช้ชื่อเดียวกับ input แต่เปลี่ยนนามสกุล)
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.rsplit('.', 1)[0] + '.xlsx'

    # แปลงไฟล์
    try:
        HTMLToExcelConverter.convert_file(input_file, output_file)
    except Exception as e:
        print(f"ไม่สามารถแปลงไฟล์ได้: {str(e)}")
        sys.exit(1) 